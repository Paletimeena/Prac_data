from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import time
import automate
import time
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
"""

#List1=["S.NO.","Test_Case_Id","Expected_Result","Actual_Result","Status","Date & Time"]
#List2=["Testcase_Passed","Testcase_Fail"]
def Heading():	
	font_color=Font(color=colors.BLUE)
	for i, statN in enumerate(List1):
		ws1.cell(row=1, column=i+1).font = font_color
		ws1.cell(row=1, column=i+1).value = statN
	for i, statN in enumerate(List2):
		ws2.cell(row=1, column=i+1).font = font_color
		ws2.cell(row=1, column=i+1).value = statN
"""

def callxl():

	global wb,ws1,ws2
	wb = load_workbook('Sample Data.xlsx')
	ws1 = ws2 = wb.active
	ws1 = wb.get_sheet_by_name("Sheet2")
"""
	#ws2 = wb.get_sheet_by_name("Status")
	row1=int(list1[1])
	col=int(list1[2])
	Heading()
	ws1.cell(row=row1,column=col-1).value=row1-2             #for Serial number
	ws1.cell(row=row1,column=col).value=list1[0]             #for testcase id 
	ws1.cell(row=row1,column=col+1).value=list1[3]           #for expected result
	ws1.cell(row=row1,column=col+2).value=list1[4]           #for actual result
	if(list1[3]==list1[4]):
		status="Pass"
		font_colors=Font(color=colors.GREEN)
	else:
		status="Fail"
		font_colors=Font(color=colors.RED)
	
	ws1.cell(row=row1,column=col+3).font=font_colors	
	ws1.cell(row=row1,column=col+3).value=status           #for status
	ws1.cell(row=row1,column=col+4).value=str(time.asctime(time.localtime(time.time()))) #for time & date	
	wb.save('Sample.xlsx')
	check_count()
	wb.close()
	
def check_count():
	#pass_count = 0	
	#fail_count = 0
	max_row_count = ws1.max_row
	for row in ws1.iter_rows(min_row=2, min_col=2, max_row=max_row_count, max_col=5):
		for cell in row:        		
			if cell.value != None :	
				if "Pass" in cell.value:	
					pass_count +=1
				elif "Fail" in cell.value:
					fail_count +=1
	ws2.cell(row=2,column=1).value=pass_count
	ws2.cell(row=2,column=2).value=fail_count
	wb.save('Sample.xlsx')
"""


def get_link():
	max_row_count = ws1.max_row
	max_col_count = ws1.max_column
	print max_row_count
	print max_col_count
	for row in ws1.iter_rows(min_row=2, min_col=1, max_row=max_row_count, max_col=max_col_count):
		print "checking for rows\n"
		for cell in row:
			print cell.value,cell.row,cell.column
			print type(cell.value)
			automate.call_link(cell.value)

callxl()
get_link()







	
